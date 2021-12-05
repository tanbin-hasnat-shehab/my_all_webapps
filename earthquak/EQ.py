import streamlit as st
from openpyxl import load_workbook
import pandas as pd
def Earthquak_load_cal():
	
	wb=load_workbook('earthquak/db.xlsx')
	
	sheet=wb['sheet']

	st.header("Earthquak load calculation as per BNBC'20")
	loc_arr=[]
	for i in range(10,76):
		loc_arr.append(sheet.cell(row=i,column=1).value)
	location=st.selectbox('Location',loc_arr)
	Z=float(sheet.cell(row=loc_arr.index(location)+10, column=2).value)
	print(f'z is {Z}')
	if Z==.12:
		zone=0
	if Z==.20:
		zone=1
	if Z==.28:
		zone=2
	if Z==.36:
		zone=3
	print(f'zone={zone}')
	num_o_sto=st.checkbox('input number of storey or input height of structure')
	if num_o_sto:
		basement_height=st.text_input('basement height in meter')
		number_of_storey=st.text_input('number of storey without basement')
		each_storey_height=st.text_input('height of each storey in meter')
		
		if each_storey_height:
			total_height=float(basement_height)+int(number_of_storey)*float(each_storey_height)
			print(total_height)
	else:
		total_height=st.text_input('height of the structure in meter')
		if total_height:
			total_height=float(total_height)
			print('this from else',total_height)


	occupency_arr=[]
	for i in range(161,164):
		occupency_arr.append(sheet.cell(row=i,column=1).value)

	occupency_catagory=st.selectbox('Occupency catagory',occupency_arr)
	if occupency_catagory:
		print(occupency_catagory)
		importance_factor=float(sheet.cell(row=occupency_arr.index(occupency_catagory)+161, column=2).value)
	
	####


	ita=1
	st.title('Soil report needed for N values')
	pen_spacing=st.text_input('equi-depth in SPT per penetration','5')
	if pen_spacing:
		pen_spacing=float(pen_spacing)
	pen_val_inp=st.text_input('penetration values in feet separated by a comma')
	if pen_val_inp:
		pen_in_str=pen_val_inp.split(',')
		penetration_val=[]
		for val in pen_in_str:
			penetration_val.append(float(val))
		sum_denomi=0
		for pen in penetration_val:
			sum_denomi=sum_denomi+(pen_spacing)/pen
		sum_nomi=0
		for i in range(len(penetration_val)):
			sum_nomi=sum_nomi+pen_spacing
		N=sum_nomi/sum_denomi
		if N<15:
			site_class='SD'
			site_class_no=3
			seismic_deg_cat_row=178
		elif N>=15 and N<=50:
			site_class='SC'
			seismic_deg_cat_row=177
			site_class_no=2
		elif N>50 and N<100:
			site_class='SB'
			site_class_no=1
			seismic_deg_cat_row=176
		else:
			site_class='SA'
			seismic_deg_cat_row=175
			site_class_no=0

		if occupency_catagory=='IV':
			seismic_deg_cat_col_start=6
		else:
			seismic_deg_cat_col_start=2

		seismic_deg_cat=sheet.cell(row=seismic_deg_cat_row,column=seismic_deg_cat_col_start+zone).value
		print(seismic_deg_cat)
		if seismic_deg_cat=='B':
			seismic_force_resisting_system_suggestion='ordinary seismic_force_resisting_system'
		if seismic_deg_cat=='C':
			seismic_force_resisting_system_suggestion='intermediate seismic_force_resisting_system'
		if seismic_deg_cat=='D':
			seismic_force_resisting_system_suggestion='special seismic_force_resisting_system'
		print(seismic_force_resisting_system_suggestion)
		my_splsh=f'<p style="font-family:Courier; color:Blue; font-size: 20px;">{seismic_force_resisting_system_suggestion}</p>'
		st.markdown(my_splsh, unsafe_allow_html=True)


		seismic_force_resisting_system_arr=[]
		for i in range(189,222):
			seismic_force_resisting_system_arr.append(sheet.cell(row=i,column=1).value)
		seismic_force_resisting_system=st.selectbox('Seismic Force-Resisting System',seismic_force_resisting_system_arr)
		if seismic_force_resisting_system:
			R=float(sheet.cell(row=seismic_force_resisting_system_arr.index(seismic_force_resisting_system)+189, column=2).value)
			system_over_strength_factor=float(sheet.cell(row=seismic_force_resisting_system_arr.index(seismic_force_resisting_system)+189, column=3).value)
			Deflection_Amplification_Factor=float(sheet.cell(row=seismic_force_resisting_system_arr.index(seismic_force_resisting_system)+189, column=4).value)
			Ss=float(sheet.cell(row=258,column=2+zone).value)
			S1=float(sheet.cell(row=259,column=2+zone).value)
			Fa=float(sheet.cell(row=264+site_class_no,column=2+zone).value)
			Fv=float(sheet.cell(row=273+site_class_no,column=2+zone).value)
			
		




			








		for i in range(150,154):
			if sheet.cell(row=i, column=1).value==site_class:
				S=float(sheet.cell(row=i, column=2).value)
				Tb=float(sheet.cell(row=i, column=3).value)
				Tc=float(sheet.cell(row=i, column=4).value)
				Td=float(sheet.cell(row=i, column=5).value)
		#cal of T
		st_type_arr=[]
		for i in range(242,246):
			st_type_arr.append(sheet.cell(row=i,column=1).value)
		st_type=st.selectbox('structure type',st_type_arr)
		if st_type:
			G=float(sheet.cell(row=st_type_arr.index(st_type)+242, column=2).value)
			m=float(sheet.cell(row=st_type_arr.index(st_type)+242, column=3).value)
			T=G*total_height**m
			print(f'G={G} and m={m} N={N} site class={site_class} S={S} Tb={Tb} R={R}')

			if T>=0 and T<=Tb:
				Cs=S*(1+T/Tb*(2.5*ita-1))
			elif T>=Tb and T<=Tc:
				Cs=2.5*S*ita
			elif T>=Tc and T<=Td:
				Cs=2.5*S*ita*(Tc/T)
			elif T>=Td and T<=4:
				Cs=2.5*S*ita*(Tc*Td)/(T**2)
			else:
				pass
			Sa=2/3*Z*importance_factor*Cs/R
			Sa_min=0.67*.11*Z*importance_factor*S
			if Sa<Sa_min:
				Sa=Sa_min
			st.title(f'Base shear is {Sa*100}%')




			to_df2=[]
			
			to_df2.append(['Ecc. ratio (All depth)',0.05])
			to_df2.append(['Ct(ft).x =',m])
			to_df2.append(['Response modification factor,R',R])
			to_df2.append(['system over strength,Omega',system_over_strength_factor])
			to_df2.append(['Deflection amplification ,Cd',Deflection_Amplification_Factor])
			to_df2.append(['Occupency Importance,I',importance_factor])
			to_df2.append(['0.2 sec spectral Accel,Ss',Ss])
			to_df2.append(['1 sec spectral Accel,S1',S1])
			to_df2.append(['Long period transition period',2])
			to_df2.append(['Site coeffecient,Fa',Fa])
			to_df2.append(['Site coeffecient,Fv',Fv])
			to_df2.append(['zone factor,Z',Z])
			to_df2.append(['Time period,T',T])

			to_df2.append(['Site class',site_class])
			to_df2.append(['Seismic design catagory',seismic_deg_cat])
			to_df2.append(['Seismic force resisting system',seismic_force_resisting_system])
			#to_df2.append(['',])
			df2=pd.DataFrame(to_df2,columns=['Name of coeffecient','value'])
			st.subheader('Co coeffecients needed for etabs 2016 or higher')
			st.write(df2)





			siesmic_weight=st.text_input('seismic weight from stadd or etabbs')
			if siesmic_weight:
				base_shear=float(siesmic_weight)*Sa
				st.title(f'Base shear load is {base_shear}')
				floor_loads_cal=st.checkbox('calculate each floor loads')
				if floor_loads_cal:
					n=st.text_input('no of heights')

					c1,c2=st.beta_columns((1,1))
					if n:
						n=int(n)
						with c1:
							st.subheader('weight of each storey')
							task_names={}
							task_names_one_d=[]
							for i in range(n):
							    task_names[f'weight_of_storey{i}']=''
							for k, v in task_names.items():
								task_names[k] = st.text_input(k, v)
								#st.write(task_names[k])
							
						with c2:
							st.subheader('h')
							durations={}
							duration_one_d=[]
							for i in range(n):
								durations[f'height_no{i}']=''
							for k, v in durations.items():
								durations[k] = st.text_input(k, v)
								#st.write(durations[k])
							
						
						

						proc_btn=st.button('show base shear distribution')

						if proc_btn:

							for i in range(n):
								task_names_one_d.append(float(task_names[f'weight_of_storey{i}']))
							print(task_names_one_d)
							floor_loads_arr=task_names_one_d
							
							for i in range(n):
								duration_one_d.append(float(durations[f'height_no{i}']))
							print(duration_one_d)
							heights_arr=duration_one_d




							denomi=0
							#k
							if T<=0.5:
								k=1
							elif T>0.5 and T<2.5:
								k=(T-.5)*(-1)/(.5-2.5)+1
							else:
								k=2
							for i in range(len(heights_arr)):
								denomi=denomi+floor_loads_arr[i]*heights_arr[i]**k
							result_floor_loads=[]
							for i in range(len(heights_arr)):
								result_floor_loads.append(base_shear*floor_loads_arr[i]*heights_arr[i]**k/denomi)
							to_df=[]
							for i in range(len(heights_arr)):
								to_df.append([result_floor_loads[i],heights_arr[i]])
							df=pd.DataFrame(to_df,columns=['Floor Loads','Height'])

							st.write(df)
							st.subheader('reference')
							st.video('https://www.youtube.com/watch?v=gMHAYcGOAmg&list=LL&index=1')


	
								


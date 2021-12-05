import streamlit as st
import codecs
import streamlit.components.v1 as stc
from anastruct import SystemElements
import math
import matplotlib.pyplot as plt
from PIL import Image
from openpyxl import load_workbook,Workbook
import json
def sfd_bmd():
	wb=Workbook()
	  
	wb.create_sheet("DISPLACEMENTS")
	wb.create_sheet('BENDING_MOMENT')
	wb.create_sheet('AXIAL_FORCE')
	wb.create_sheet('SHEAR_FORCE')
	wb.create_sheet('REACTIONS')
	wb.save('sfd/data.xlsx')
	wt=load_workbook('sfd/b.xlsx')
	dsheet=wt['sheet']


	#st.set_page_config(layout="wide")
	#st.title(f'the value is {dsheet.cell(row=5,column=2).value}')
	js_str=st.text_input('paste here')
	aaa=st.slider("zoom in out results", 0, 300, value=30)
	if st.button("Run"):
		
		#js_string = cookie_manager.get(cookie)
		#jtopy=json.dumps(js_str)
		value=json.loads(js_str)

		
		

		#print(value)
			



		ss = SystemElements()

		new_data=[]
		for i in range(len(value)):
			if value[i] not  in new_data:
				new_data.append(value[i])


		line_arr=[]
		for data in new_data:
			if data['type']=='line':
				line_arr.append(new_data.index(data))
			if data['type']=='fixed_support' or data['type']=='hinged_support' or data['type']=='rolled_support':
				line_arr.append(new_data.index(data))
				break


		gr_line=[]



		print(f'thi9s is line arr     {line_arr}')



		for data in new_data:
			if data['type']=='type_of_st':
				structure_type=data['value']
			if data['type']=='grid_spacing':
				grid=data['value']
		for data in new_data:
			if data['type']=='line':
				data['x1']=data['x1']/grid
				data['y1']=data['y1']/grid
				data['x2']=data['x2']/grid
				data['y2']=data['y2']/grid
			if data['type']=='fixed_support':
				data['x']=data['x']/grid
				data['y']=data['y']/grid
				
			if data['type']=='hinged_support' or data['type']=='rolled_support':
				data['x']=data['x']/grid
				data['y']=data['y']/grid
				
			if data['type']=='point_load':
				data['pos_x']=data['pos_x']/grid
				data['pos_y']=data['pos_y']/grid


		my_points_x=[]
		my_points_y=[]
		joint_points=[]

		for data in new_data:
			if data['type']=='line':
				joint_points.append((data['x1'],data['y1']))
				joint_points.append((data['x2'],data['y2']))



				my_points_x.append(data['x1'])
				my_points_y.append(data['y1'])
				my_points_x.append(data['x2'])
				my_points_y.append(data['y2'])

			
			if data['type']=='fixed_support':
				my_points_x.append(data['x'])
				my_points_y.append(data['y'])

				
			if data['type']=='hinged_support' or data['type']=='rolled_support':
				my_points_x.append(data['x'])
				my_points_y.append(data['y'])

				
			if data['type']=='point_load':
				my_points_x.append(data['pos_x'])
				my_points_y.append(data['pos_y'])

				
				
				

					

			
		#print(new_data)
		
		no_of_member=0
		members=[]
		no_of_load=0
		loads=[]
		no_of_fixed=0
		fixed=[]
		no_of_hinged=0
		hinged=[]
		no_of_rolled=0
		rolled=[]
		member_loads=[]

		for i in range(len(new_data)):
			if new_data[i]['type']=='line':
				no_of_member+=1
				members.append(new_data[i])
				member_loads.append(new_data[i]['W'])
			if new_data[i]['type']=='point_load':
				no_of_load+=1
				loads.append(new_data[i])
			if new_data[i]['type']=='fixed_support':
				no_of_fixed+=1
				fixed.append(new_data[i])
			if new_data[i]['type']=='hinged_support':
				no_of_hinged+=1
				hinged.append(new_data[i])
			if new_data[i]['type']=='rolled_support':
				no_of_rolled+=1
				rolled.append(new_data[i])


		if structure_type=='truss':
			for i in range(no_of_member):
				ss.add_truss_element(location=[[members[i]['x1'], members[i]['y1']]   , [members[i]['x2'], members[i]['y2']] ],EA=5000*members[i]['EA'])    
			for i in range(no_of_load):
				ss.point_load(Fx=loads[i]['load'],rotation=loads[i]['rotation'], node_id=ss.find_node_id([loads[i]['pos_x'],loads[i]['pos_y']]))
			


		if structure_type=='frame' or structure_type=='beam':

			for i in range(len(line_arr)-1):
				ss.add_element(location=[  [new_data[line_arr[i]]['x1'], new_data[line_arr[i]]['y1']], [new_data[line_arr[i]]['x2'], new_data[line_arr[i]]['y2']] ],EA=5000*new_data[line_arr[i]]['EA'],EI=5000)   
				number_of_nodes=line_arr[i+1]-line_arr[i]-1
				if not number_of_nodes==0:
					for j in range(number_of_nodes):
						#print(f'this is test 2 {new_data[line_arr[i+j]+1]}')
						if (new_data[line_arr[i]+j+1]['pos_x'],new_data[line_arr[i]+j+1]['pos_y'] )!=(new_data[line_arr[i]]['x1'],new_data[line_arr[i]]['y1']) and (new_data[line_arr[i]+j+1]['pos_x'],new_data[line_arr[i]+j+1]['pos_y'] )!=(new_data[line_arr[i]]['x2'],new_data[line_arr[i]]['y2']):
							ss.insert_node(element_id=ss.id_last_element, location=[new_data[line_arr[i]+j+1]['pos_x'], new_data[line_arr[i]+j+1]['pos_y']])
			###################		
			cr=0
			k=0
			for i in range(len(line_arr)-1):

				print(f'limits {line_arr[i]} to {line_arr[i+1]}')
				for j in range(line_arr[i],line_arr[i+1]):
					if new_data[j+1]['type']=='point_load':
						if (new_data[j+1]['pos_x'],new_data[j+1]['pos_y']) not in joint_points:
							cr+=1
							ss.q_load(q=-new_data[line_arr[i]]['W'], element_id=cr, direction='element')
							print(f'this is {j+1} section and value is  {new_data[line_arr[i]]["W"]}')
					else:
						cr+=1
						ss.q_load(q=-new_data[line_arr[i]]['W'], element_id=cr, direction='element')




			ss.q_load(q=-member_loads[-1], element_id=ss.id_last_element, direction='element')		

			print('\n\n')
			######################
			print(joint_points)
			for i in range(no_of_load):
				ss.point_load(Fx=loads[i]['load'],rotation=loads[i]['rotation'], node_id=ss.find_node_id([loads[i]['pos_x'],loads[i]['pos_y']]))
			





		for i in range(no_of_fixed):
			ss.add_support_fixed(node_id=ss.find_node_id([fixed[i]['x'],fixed[i]['y']]))






		for i in range(no_of_hinged):
			
			ss.add_support_hinged(node_id=ss.find_node_id([hinged[i]['x'],hinged[i]['y']]))
		for i in range(no_of_rolled):
			
			ss.add_support_roll(node_id=ss.find_node_id([rolled[i]['x'],rolled[i]['y']]))
		ss.solve()

		wb=load_workbook('data.xlsx')
		dis=wb['DISPLACEMENTS']
		ben=wb['BENDING_MOMENT']
		axi=wb['AXIAL_FORCE']
		shr=wb['SHEAR_FORCE']
		rea=wb['REACTIONS']	
		

		ss.show_structure()
		plt.xlim([min(my_points_x)-aaa,max(my_points_x)+aaa])
		plt.ylim([min(my_points_y)-aaa,max(my_points_y)+aaa])
		plt.title('STRUCTURE')
		plt.savefig('sfd/my-figure1.png')
		image1 = Image.open('sfd/my-figure1.png')
		st.image(image1)
		
		ss.show_reaction_force()
		plt.xlim([min(my_points_x)-aaa,max(my_points_x)+aaa])
		plt.ylim([min(my_points_y)-aaa,max(my_points_y)+aaa])
		plt.title('REACTIONS')
		plt.savefig('sfd/my-figure2.png')
		image2 = Image.open('sfd/my-figure2.png')
		st.image(image2)
		

		ss.show_displacement()
		plt.xlim([min(my_points_x)-aaa,max(my_points_x)+aaa])
		plt.ylim([min(my_points_y)-aaa,max(my_points_y)+aaa])
		plt.title('DISPLACEMENTS')
		plt.savefig('sfd/my-figure22.png')
		image22 = Image.open('sfd/my-figure22.png')
		st.image(image22)
		m=ss.show_displacement(show=False,values_only=True)
		for i in range(0,len(m[0])):
			dis.cell(row=i+1,column=1).value=m[0][i]
			dis.cell(row=i+1,column=2).value=m[1][i]
		
		



		ss.show_axial_force()
		plt.xlim([min(my_points_x)-aaa,max(my_points_x)+aaa])
		plt.ylim([min(my_points_y)-aaa,max(my_points_y)+aaa])
		plt.title('AXIAL FORCE DIAGRAM')
		plt.savefig('sfd/my-figure3.png')
		image3 = Image.open('sfd/my-figure3.png')
		st.image(image3)
		m=ss.show_axial_force(show=False,values_only=True)
		for i in range(0,len(m[0])):
			axi.cell(row=i+1,column=1).value=m[0][i]
			axi.cell(row=i+1,column=2).value=m[1][i]
		if structure_type=='truss':
			st.title('Combined image')

		if structure_type=='beam' or structure_type=='frame':
			ss.show_bending_moment()
			plt.xlim([min(my_points_x)-aaa,max(my_points_x)+aaa])
			plt.ylim([min(my_points_y)-aaa,max(my_points_y)+aaa])
			plt.title('BENDING MOMENT DIAGRAM')
			plt.savefig('sfd/my-figure4.png')
			image4 = Image.open('sfd/my-figure4.png')
			st.image(image4)
			m=ss.show_bending_moment(show=False,values_only=True)
			for i in range(0,len(m[0])):
				ben.cell(row=i+1,column=1).value=m[0][i]
				ben.cell(row=i+1,column=2).value=m[1][i]

			ss.show_shear_force()
			plt.xlim([min(my_points_x)-aaa,max(my_points_x)+aaa])
			plt.ylim([min(my_points_y)-aaa,max(my_points_y)+aaa])
			plt.title('SHEAR FORCE DIAGRAM')
			plt.savefig('sfd/my-figure5.png')
			image5 = Image.open('sfd/my-figure5.png')
			st.image(image5)
			m=ss.show_shear_force(show=False,values_only=True)
			for i in range(0,len(m[0])):
				shr.cell(row=i+1,column=1).value=m[0][i]
				shr.cell(row=i+1,column=2).value=m[1][i]
			st.title('Combined image')
			im1=Image.open('sfd/my-figure1.png')
			im2=Image.open('sfd/my-figure2.png')
			im22=Image.open('sfd/my-figure22.png')
			im3=Image.open('sfd/my-figure3.png')
			im4=Image.open('sfd/my-figure4.png')
			im5=Image.open('sfd/my-figure5.png')

			im_size=im1.size
			combined_im=Image.new('RGB', (im_size[0],im_size[1]*6),(250,250,250))
			combined_im.paste(im1,(0,0))
			combined_im.paste(im2,(0,im_size[1]*1))
			combined_im.paste(im22,(0,im_size[1]*2))
			combined_im.paste(im3,(0,im_size[1]*3))
			combined_im.paste(im4,(0,im_size[1]*4))
			combined_im.paste(im5,(0,im_size[1]*5))
			combined_im.save('comb.png','PNG')
			st.image(combined_im)
		else:
			im1=Image.open('sfd/my-figure1.png')
			im2=Image.open('sfd/my-figure2.png')
			im22=Image.open('sfd/my-figure22.png')
			im3=Image.open('sfd/my-figure3.png')
			

			im_size=im1.size
			combined_im=Image.new('RGB', (im_size[0],im_size[1]*4),(250,250,250))
			combined_im.paste(im1,(0,0))
			combined_im.paste(im2,(0,im_size[1]*1))
			combined_im.paste(im22,(0,im_size[1]*2))
			combined_im.paste(im3,(0,im_size[1]*3))
			
			combined_im.save('sfd/comb.png','PNG')
			st.image(combined_im)


		
		wb.save('sfd/data.xlsx')
		

	def my_html(html_file,width=2000,height=2000):
		html_file=codecs.open(html_file,'r')
		page=html_file.read()
		stc.html(page,width=width,height=height,scrolling=False)


	my_html('sfd/a.html')






